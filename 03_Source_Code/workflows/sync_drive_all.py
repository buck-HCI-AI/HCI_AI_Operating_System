"""
sync_drive_all.py — Google Drive full-read ingestion, ALL Shared Drives

Found 2026-07-07: sync_drive.py only ever walked a local filesystem mirror of
"My Drive/HCI AI - Master" (Buck's Google Drive Desktop mount). It never saw
any of the 18 real per-project Shared Drives (574 Johnson Drive, 813
McSkimming Spitzley, 1355 Riverside Ryan's, 101 Francis, 64 Eastwood, and 13
more) - the actual vendor bids, budgets, permits, and contracts. This version
walks every Shared Drive via the API directly, so it works regardless of
whether any particular machine has Drive Desktop mounted, and actually
reaches the real project source documents.

Same target as sync_drive.py: Qdrant `drive_memory` collection, Postgres
`drive_sync_log` (now keyed by Drive file ID instead of local path so both
scripts can coexist without colliding).
"""
import sys, os, re, json, ssl, io, urllib.request, urllib.parse
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "integrations"))

from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), "..", "..", ".env"))

import certifi
import psycopg2
from qdrant_client import QdrantClient
from qdrant_client.models import PointStruct, Distance, VectorParams
from fastembed import TextEmbedding
from credentials import get_google_token

SSL_CTX = ssl.create_default_context(cafile=certifi.where())
BASE_URL = "https://www.googleapis.com/drive/v3"
CHUNK_SIZE = 800
CHUNK_OVERLAP = 100
QDRANT_COLLECTION = "drive_memory"
VECTOR_DIM = 384
ID_OFFSET = 40000

DB = dict(
    host=os.environ.get("POSTGRES_HOST", "localhost"),
    port=int(os.environ.get("POSTGRES_PORT", 5432)),
    dbname=os.environ.get("POSTGRES_DB", "hci_os"),
    user=os.environ.get("POSTGRES_USER", "hci_admin"),
    password=os.environ.get("POSTGRES_PASSWORD", ""),
)

_PROCESSABLE_MIMES = {
    "application/pdf": "pdf",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "docx",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
    "application/vnd.google-apps.document": "gdoc",
    "application/vnd.google-apps.spreadsheet": "gsheet",
}
FILE_FIELDS = "id,name,mimeType,modifiedTime,size,driveId"

_embedder = None


def get_embedder():
    global _embedder
    if _embedder is None:
        _embedder = TextEmbedding("BAAI/bge-small-en-v1.5")
    return _embedder


def embed(texts):
    return [list(v) for v in get_embedder().embed(texts)]


def db_conn():
    return psycopg2.connect(**DB)


def ensure_table(conn):
    with conn.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS drive_sync_log_v2 (
                id          SERIAL PRIMARY KEY,
                file_id     TEXT UNIQUE NOT NULL,
                file_name   TEXT,
                file_type   TEXT,
                drive_name  TEXT,
                modified    TEXT,
                chunks      INT DEFAULT 0,
                synced_at   TIMESTAMPTZ DEFAULT NOW()
            )
        """)
    conn.commit()


def get_synced(conn) -> dict:
    with conn.cursor() as cur:
        cur.execute("SELECT file_id, modified FROM drive_sync_log_v2")
        return {row[0]: row[1] for row in cur.fetchall()}


def upsert_log(conn, file_id, name, ftype, drive_name, modified, chunks):
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO drive_sync_log_v2 (file_id, file_name, file_type, drive_name, modified, chunks)
            VALUES (%s,%s,%s,%s,%s,%s)
            ON CONFLICT (file_id) DO UPDATE SET
                modified=EXCLUDED.modified, chunks=EXCLUDED.chunks, synced_at=NOW()
        """, (file_id, name, ftype, drive_name, modified, chunks))
    conn.commit()


def get_qdrant():
    host = os.environ.get("QDRANT_HOST", "localhost")
    port = int(os.environ.get("QDRANT_PORT", 6333))
    client = QdrantClient(host=host, port=port)
    try:
        client.get_collection(QDRANT_COLLECTION)
    except Exception:
        client.create_collection(
            collection_name=QDRANT_COLLECTION,
            vectors_config=VectorParams(size=VECTOR_DIM, distance=Distance.COSINE),
        )
    return client


def ingest_chunks(qdrant, chunks, start_id):
    if not chunks:
        return start_id
    texts = [c["text"] for c in chunks]
    vectors = embed(texts)
    points = [PointStruct(id=start_id + i, vector=vec, payload=chunk)
              for i, (chunk, vec) in enumerate(zip(chunks, vectors))]
    qdrant.upsert(collection_name=QDRANT_COLLECTION, points=points)
    return start_id + len(points)


def chunk_text(text, source, file_name, file_type, drive_name):
    text = re.sub(r"\s{3,}", "\n\n", text.strip())
    chunks = []
    pos = 0
    while pos < len(text):
        end = pos + CHUNK_SIZE
        chunk = text[pos:end]
        if chunk.strip():
            chunks.append({
                "text": chunk, "source": source, "file_name": file_name,
                "file_type": file_type, "drive_name": drive_name,
            })
        pos += CHUNK_SIZE - CHUNK_OVERLAP
    return chunks


def _get_json(url, params=None):
    if params:
        url += "?" + urllib.parse.urlencode(params)
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {get_google_token('drive')}"})
    with urllib.request.urlopen(req, context=SSL_CTX, timeout=30) as r:
        return json.loads(r.read())


def _get_bytes(url, params=None):
    if params:
        url += "?" + urllib.parse.urlencode(params)
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {get_google_token('drive')}"})
    with urllib.request.urlopen(req, context=SSL_CTX, timeout=60) as r:
        return r.read()


def list_shared_drives():
    return _get_json(f"{BASE_URL}/drives", {"pageSize": 100}).get("drives", [])


def walk_drive_files(drive_id: str, max_files: int = 2000) -> list:
    """All processable files anywhere in a Shared Drive, any depth, via one query
    (Drive API search is flat across the whole drive when scoped by driveId -
    no need to recurse folder by folder). max_files is a real hard cap - found
    2026-07-07 the original version only checked the cap between full 200-file
    pages, so a small test limit could still process 10x+ more than requested
    and blow through memory on a large PDF batch."""
    files, token = [], None
    mime_clause = " or ".join(f"mimeType='{m}'" for m in _PROCESSABLE_MIMES)
    while len(files) < max_files:
        params = {
            "q": f"trashed=false and ({mime_clause})",
            "fields": f"files({FILE_FIELDS}),nextPageToken",
            "pageSize": min(200, max_files - len(files)),
            "driveId": drive_id,
            "corpora": "drive",
            "supportsAllDrives": "true",
            "includeItemsFromAllDrives": "true",
        }
        if token:
            params["pageToken"] = token
        resp = _get_json(f"{BASE_URL}/files", params)
        files.extend(resp.get("files", []))
        token = resp.get("nextPageToken")
        if not token:
            break
    return files[:max_files]


def extract_pdf_bytes(data: bytes) -> str:
    import pdfplumber
    pages = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                pages.append(t)
    return "\n\n".join(pages)


def extract_docx_bytes(data: bytes) -> str:
    from docx import Document
    doc = Document(io.BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def extract_xlsx_bytes(data: bytes) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    parts = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None and str(c).strip()]
            if cells:
                rows.append("  ".join(cells))
        if rows:
            parts.append(f"[Sheet: {sheet}]\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(parts)


def extract_google_export(file_id: str, mime: str) -> str:
    data = _get_bytes(f"{BASE_URL}/files/{file_id}/export", {"mimeType": mime})
    return data.decode("utf-8", errors="replace")


def run(max_files_per_drive: int = 300, drive_name_filter: str = None) -> dict:
    print("\n=== Drive Sync (all Shared Drives) ===")
    conn = db_conn()
    ensure_table(conn)
    qdrant = get_qdrant()
    synced = get_synced(conn)

    drives = list_shared_drives()
    if drive_name_filter:
        drives = [d for d in drives if drive_name_filter.lower() in d["name"].lower()]
    print(f"  {len(drives)} Shared Drive(s) to scan")

    try:
        info = qdrant.get_collection(QDRANT_COLLECTION)
        next_id = ID_OFFSET + max(info.points_count, 0)
    except Exception:
        next_id = ID_OFFSET

    totals = {"drives_scanned": 0, "files_found": 0, "files_synced": 0,
              "files_skipped": 0, "errors": 0, "total_chunks": 0, "per_drive": {}}

    for drive in drives:
        drive_id, drive_name = drive["id"], drive["name"]
        try:
            files = walk_drive_files(drive_id, max_files_per_drive)
        except Exception as e:
            print(f"  ✗ {drive_name}: list failed - {e}")
            continue
        totals["drives_scanned"] += 1
        totals["files_found"] += len(files)
        print(f"\n  --- {drive_name} ({len(files)} processable files) ---")
        drive_new = 0

        for f in files:
            fid, fname, mime = f["id"], f["name"], f["mimeType"]
            modified = f.get("modifiedTime", "")
            ftype = _PROCESSABLE_MIMES.get(mime, "other")

            if fid in synced and synced[fid] == modified:
                totals["files_skipped"] += 1
                continue

            # Skip huge binary files - pdfplumber on a 100+ page scanned PDF can
            # use several GB of RAM. 30MB covers real construction docs comfortably.
            size_bytes = f.get("size")
            if size_bytes and int(size_bytes) > 30 * 1024 * 1024:
                print(f"    - skip (too large, {int(size_bytes)//1024//1024}MB): {fname[:60]}")
                totals["files_skipped"] += 1
                continue

            try:
                if ftype == "pdf":
                    text = extract_pdf_bytes(_get_bytes(f"{BASE_URL}/files/{fid}", {"alt": "media", "supportsAllDrives": "true"}))
                elif ftype == "docx":
                    text = extract_docx_bytes(_get_bytes(f"{BASE_URL}/files/{fid}", {"alt": "media", "supportsAllDrives": "true"}))
                elif ftype == "xlsx":
                    text = extract_xlsx_bytes(_get_bytes(f"{BASE_URL}/files/{fid}", {"alt": "media", "supportsAllDrives": "true"}))
                elif ftype == "gdoc":
                    text = extract_google_export(fid, "text/plain")
                elif ftype == "gsheet":
                    text = extract_google_export(fid, "text/tab-separated-values")
                else:
                    continue

                if not text or not text.strip():
                    totals["errors"] += 1
                    continue

                chunks = chunk_text(text, f"{drive_name}/{fname}", fname, ftype, drive_name)
                if chunks:
                    next_id = ingest_chunks(qdrant, chunks, next_id)
                    totals["total_chunks"] += len(chunks)

                upsert_log(conn, fid, fname, ftype, drive_name, modified, len(chunks))
                totals["files_synced"] += 1
                drive_new += 1
                print(f"    ✓ {fname[:60]} ({len(chunks)} chunks)")

            except Exception as e:
                print(f"    ✗ {fname[:60]}: {e}")
                totals["errors"] += 1

        totals["per_drive"][drive_name] = {"found": len(files), "new": drive_new}

    conn.close()
    print(f"\n  ✓ Complete: {totals['files_synced']} synced, {totals['files_skipped']} unchanged, "
          f"{totals['errors']} errors, {totals['total_chunks']} chunks, across {totals['drives_scanned']} drives")
    return totals


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("--max-per-drive", type=int, default=300)
    p.add_argument("--drive", type=str, default=None)
    args = p.parse_args()
    result = run(args.max_per_drive, args.drive)
    print(json.dumps(result, indent=2))
