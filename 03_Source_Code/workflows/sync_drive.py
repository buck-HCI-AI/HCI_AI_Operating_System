"""
sync_drive.py — Google Drive full-read ingestion
Reads all HCI Drive files into Qdrant `drive_memory` collection.

Supported formats:
  PDF     → pdfplumber
  DOCX    → python-docx
  XLSX    → openpyxl (each sheet as text)
  GDOC    → Google Drive export API (text/plain), file ID from stub JSON
  GSHEET  → Google Drive export API (text/csv), file ID from stub JSON

Tracks processed files in Postgres `drive_sync_log` (path + mtime) so
re-runs only re-process files that changed.
"""
import sys, os, json, re, ssl, urllib.request, urllib.parse, hashlib
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "integrations"))

from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), "..", "..", ".env"))

import certifi
import psycopg2
from qdrant_client import QdrantClient
from qdrant_client.models import PointStruct, Distance, VectorParams
from fastembed import TextEmbedding
from credentials import get_google_token

SSL_CTX = ssl.create_default_context(cafile=certifi.where())

DRIVE_ROOT = "/Users/buckadams/Library/CloudStorage/GoogleDrive-buck@hendricksoninc.com/My Drive/HCI AI - Master "
CHUNK_SIZE = 800
CHUNK_OVERLAP = 100
QDRANT_COLLECTION = "drive_memory"
VECTOR_DIM = 384
ID_OFFSET = 40000

DB = dict(
    host=os.environ.get("POSTGRES_HOST", "localhost"),
    port=int(os.environ.get("POSTGRES_PORT", 5432)),
    dbname=os.environ.get("POSTGRES_DB", "hci_os"),
    user=os.environ.get("POSTGRES_USER", "hci_admin"),
    password=os.environ.get("POSTGRES_PASSWORD", ""),
)
SKIP_DIRS = {"_archive", "__pycache__", ".git"}
SKIP_EXTS = {".py", ".sh", ".json", ".yaml", ".yml", ".gitignore", ".env", ".log", ".sqlite", ".db"}

_embedder = None

def get_embedder():
    global _embedder
    if _embedder is None:
        _embedder = TextEmbedding("BAAI/bge-small-en-v1.5")
    return _embedder


def embed(texts: list[str]) -> list[list[float]]:
    return [list(v) for v in get_embedder().embed(texts)]


# ── Postgres ──────────────────────────────────────────────────────────────────

def db_conn():
    return psycopg2.connect(**DB)


def ensure_table(conn):
    with conn.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS drive_sync_log (
                id          SERIAL PRIMARY KEY,
                file_path   TEXT UNIQUE NOT NULL,
                file_name   TEXT,
                file_type   TEXT,
                mtime       FLOAT,
                chunks      INT DEFAULT 0,
                synced_at   TIMESTAMPTZ DEFAULT NOW()
            )
        """)
    conn.commit()


def get_synced(conn) -> dict:
    """Returns {file_path: mtime} for all already-synced files."""
    with conn.cursor() as cur:
        cur.execute("SELECT file_path, mtime FROM drive_sync_log")
        return {row[0]: row[1] for row in cur.fetchall()}


def upsert_log(conn, path: str, name: str, ftype: str, mtime: float, chunks: int):
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO drive_sync_log (file_path, file_name, file_type, mtime, chunks)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT (file_path) DO UPDATE SET
                mtime=EXCLUDED.mtime, chunks=EXCLUDED.chunks, synced_at=NOW()
        """, (path, name, ftype, mtime, chunks))
    conn.commit()


# ── Qdrant ────────────────────────────────────────────────────────────────────

def get_qdrant():
    host = os.environ.get("QDRANT_HOST", "localhost")
    port = int(os.environ.get("QDRANT_PORT", 6333))
    client = QdrantClient(host=host, port=port)
    try:
        client.get_collection(QDRANT_COLLECTION)
    except Exception:
        client.create_collection(
            collection_name=QDRANT_COLLECTION,
            vectors_config=VectorParams(size=VECTOR_DIM, distance=Distance.COSINE),
        )
    return client


def ingest_chunks(qdrant, chunks: list[dict], start_id: int) -> int:
    """Embed and upsert chunks to Qdrant. Returns next available id."""
    if not chunks:
        return start_id
    texts = [c["text"] for c in chunks]
    vectors = embed(texts)
    points = []
    for i, (chunk, vec) in enumerate(zip(chunks, vectors)):
        points.append(PointStruct(
            id=start_id + i,
            vector=vec,
            payload=chunk,
        ))
    qdrant.upsert(collection_name=QDRANT_COLLECTION, points=points)
    return start_id + len(points)


# ── Text chunking ─────────────────────────────────────────────────────────────

def chunk_text(text: str, source: str, file_name: str, file_type: str) -> list[dict]:
    text = re.sub(r"\s{3,}", "\n\n", text.strip())
    chunks = []
    pos = 0
    while pos < len(text):
        end = pos + CHUNK_SIZE
        chunk = text[pos:end]
        if chunk.strip():
            chunks.append({
                "text":      chunk,
                "source":    source,
                "file_name": file_name,
                "file_type": file_type,
            })
        pos += CHUNK_SIZE - CHUNK_OVERLAP
    return chunks


# ── File extractors ───────────────────────────────────────────────────────────

def extract_pdf(path: str) -> str:
    import pdfplumber
    pages = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                pages.append(t)
    return "\n\n".join(pages)


def extract_docx(path: str) -> str:
    from docx import Document
    doc = Document(path)
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def extract_xlsx(path: str) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    parts = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None and str(c).strip()]
            if cells:
                rows.append("  ".join(cells))
        if rows:
            parts.append(f"[Sheet: {sheet}]\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(parts)


def extract_google_doc(doc_id: str, mime: str = "text/plain") -> str:
    """Export a Google Doc or Sheet as plain text via Drive API."""
    try:
        token = get_google_token("drive")
        url = (f"https://www.googleapis.com/drive/v3/files/{doc_id}/export"
               f"?mimeType={urllib.parse.quote(mime)}")
        req = urllib.request.Request(url)
        req.add_header("Authorization", f"Bearer {token}")
        with urllib.request.urlopen(req, context=SSL_CTX, timeout=30) as r:
            raw = r.read()
            return raw.decode("utf-8", errors="replace")
    except Exception as e:
        return f"[export failed: {e}]"


def read_stub_id(path: str) -> str | None:
    """Read the doc_id from a .gdoc/.gsheet stub file."""
    try:
        with open(path, "r") as f:
            data = json.load(f)
        return data.get("doc_id")
    except Exception:
        return None


# ── Walk Drive ────────────────────────────────────────────────────────────────

def walk_drive(root: str) -> list[tuple[str, float]]:
    """Return list of (abs_path, mtime) for all processable files."""
    results = []
    for dirpath, dirnames, filenames in os.walk(root):
        # Skip unwanted dirs in-place
        dirnames[:] = [d for d in dirnames if d not in SKIP_DIRS]
        for fname in filenames:
            if fname.startswith("."):
                continue
            ext = os.path.splitext(fname)[1].lower()
            if ext in SKIP_EXTS:
                continue
            if ext not in {".pdf", ".docx", ".xlsx", ".gdoc", ".gsheet"}:
                continue
            fpath = os.path.join(dirpath, fname)
            try:
                mtime = os.path.getmtime(fpath)
                results.append((fpath, mtime))
            except OSError:
                pass
    return results


# ── Main ──────────────────────────────────────────────────────────────────────

def run() -> dict:
    print("\n=== Drive Sync ===")
    conn = db_conn()
    ensure_table(conn)
    qdrant = get_qdrant()
    synced = get_synced(conn)

    files = walk_drive(DRIVE_ROOT)
    print(f"  Found {len(files)} files in Drive")

    # Determine next Qdrant point ID for drive_memory
    try:
        info = qdrant.get_collection(QDRANT_COLLECTION)
        next_id = ID_OFFSET + max(info.points_count, 0)
    except Exception:
        next_id = ID_OFFSET

    new_count = skip_count = error_count = total_chunks = 0

    for fpath, mtime in files:
        # Skip if unchanged
        if fpath in synced and abs(synced[fpath] - mtime) < 1.0:
            skip_count += 1
            continue

        fname = os.path.basename(fpath)
        ext   = os.path.splitext(fname)[1].lower()
        rel   = fpath.replace(DRIVE_ROOT, "").lstrip("/")

        print(f"  Processing: {rel[:70]}")

        try:
            text = ""

            if ext == ".pdf":
                text = extract_pdf(fpath)
                ftype = "pdf"

            elif ext == ".docx":
                text = extract_docx(fpath)
                ftype = "docx"

            elif ext == ".xlsx":
                text = extract_xlsx(fpath)
                ftype = "xlsx"

            elif ext == ".gdoc":
                doc_id = read_stub_id(fpath)
                if not doc_id:
                    print(f"    ✗ No doc_id in stub")
                    error_count += 1
                    continue
                text = extract_google_doc(doc_id, "text/plain")
                ftype = "gdoc"

            elif ext == ".gsheet":
                doc_id = read_stub_id(fpath)
                if not doc_id:
                    print(f"    ✗ No doc_id in stub")
                    error_count += 1
                    continue
                # Export as tab-separated (more info than CSV for multi-sheet)
                text = extract_google_doc(doc_id, "text/tab-separated-values")
                ftype = "gsheet"

            else:
                skip_count += 1
                continue

            if not text or not text.strip() or text.startswith("[export failed"):
                print(f"    ✗ No text extracted: {text[:60] if text else 'empty'}")
                error_count += 1
                continue

            chunks = chunk_text(text, rel, fname, ftype)
            if chunks:
                next_id = ingest_chunks(qdrant, chunks, next_id)
                total_chunks += len(chunks)

            upsert_log(conn, fpath, fname, ftype, mtime, len(chunks))
            new_count += 1
            print(f"    ✓ {len(chunks)} chunks ingested")

        except Exception as e:
            import traceback
            print(f"    ✗ Error: {e}")
            traceback.print_exc()
            error_count += 1

    conn.close()

    result = {
        "status":       "success",
        "files_found":  len(files),
        "files_synced": new_count,
        "files_skipped": skip_count,
        "errors":       error_count,
        "total_chunks": total_chunks,
    }
    print(f"\n  ✓ Drive sync complete: {new_count} synced, {skip_count} unchanged, {error_count} errors, {total_chunks} chunks")
    return result


if __name__ == "__main__":
    run()
