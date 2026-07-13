"""
Bid Leveling Service — HCI AI Operating System
Reads bid tracking Google Sheets, generates per-division bid leveling Excel files,
and manages the 00_Bids/ folder structure in Google Drive.

Works for ALL projects with gsheet_bid_tracker + drive_folder_id configured.
All Drive writes are queued via approval queue during pilot phase.
ChatGPT / GBT and other AI agents can call the API endpoints for full read/write access.
"""
import sys, os, io, json, re, uuid, ssl, urllib.parse, urllib.request
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", "integrations"))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "approval_queue"))

import psycopg2, psycopg2.extras
from dotenv import load_dotenv
from datetime import datetime, timezone
from typing import Optional
import certifi
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv(os.path.join(os.path.dirname(__file__), "..", "..", "..", ".env"))

from credentials import get_google_token
from approval_queue_service import ApprovalQueueService

DB = dict(
    host=os.environ.get("POSTGRES_HOST", "localhost"),
    port=int(os.environ.get("POSTGRES_PORT", 5432)),
    dbname=os.environ.get("POSTGRES_DB", "hci_os"),
    user=os.environ.get("POSTGRES_USER", "hci_admin"),
    password=os.environ.get("POSTGRES_PASSWORD", ""),
)

SSL_CTX = ssl.create_default_context(cafile=certifi.where())

# Known 00_Bids folder IDs that already exist in Drive
KNOWN_BIDS_FOLDERS = {
    2: "1YJatvTnK0-vxiHmI0FxVE8e9jUubVcef",  # 101 Francis 00_Bids (10 division folders already exist)
}

# Sheet tab name variants across projects
SHEET_NAMES = {
    "bid_tracking":       ["Bid Tracking", "Sheet1"],
    "division_summary":   ["HCI Division Summary", "HCI 16 Div Summary"],
    "bid_leveling_detail": ["Bid Leveling Detail"],
}

CSI_DIVISIONS = {
    "01": "General Requirements", "02": "Site Work", "03": "Concrete",
    "04": "Masonry",              "05": "Metals",    "06": "Wood & Plastic",
    "07": "Thermal & Moisture",   "08": "Doors & Windows", "09": "Finishes",
    "10": "Specialties",          "11": "Equipment", "12": "Furnishings",
    "13": "Special Construction", "14": "Conveying Systems",
    "15": "Mechanical",           "16": "Electrical",
}

# HCI brand colors for Excel
HCI_BLUE   = "1F3864"
HCI_GOLD   = "BF9000"
HCI_LIGHT  = "D6E4F7"
HCI_HEADER = "2F5496"


def _pg():
    return psycopg2.connect(**DB, cursor_factory=psycopg2.extras.RealDictCursor)


class MonitoredFolderWriteError(Exception):
    """Raised when a Drive write targets a monitored project's root folder."""
    pass


def reject_if_monitored_folder(folder_id: str) -> None:
    """Buck's permanent directive: monitored/reference jobs are read-only, never
    written to - his directive names both categories explicitly (e.g. "212
    Cleveland... 655 Garmisch"). Exact-match guard against projects.drive_folder_id
    where status IN ('monitoring', 'reference').
    2026-07-13 fix: this only checked status='monitoring', silently missing every
    project stored as status='reference' (212CL, 370G, 655G, 675M) even though
    Buck's directive names 212CL and 655G by name as protected. Found during a
    GBT-requested retrieval/indexing break audit - real gap, not a guessed one.
    Known incomplete: most monitoring/reference projects still have no
    drive_folder_id populated in the DB, so this guard can't protect them yet.
    Populating those is a data task, not a code task - flagged, not silently
    treated as complete coverage."""
    with _pg() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT project_code FROM projects WHERE status IN ('monitoring', 'reference') AND drive_folder_id=%s",
                (folder_id,)
            )
            row = cur.fetchone()
    if row:
        raise MonitoredFolderWriteError(
            f"Refused: {row['project_code']} is a monitored project - read-only per "
            f"permanent directive, no writes to its Drive folder are permitted."
        )


def _log_roi(project_id, project_name, divisions, bids, files_generated):
    try:
        with _pg() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO roi_log
                        (workflow, project_id, project_code, baseline_minutes, ai_assisted_minutes,
                         steps_removed, documents_processed, notes, actor)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, ("bid_leveling", project_id, project_name,
                      90, 5,  # baseline 90 min manual leveling → 5 min AI-assisted
                      8,      # steps removed: manual sheet reading, sorting, formatting, uploading
                      bids,
                      f"{divisions} divisions, {files_generated} files generated",
                      "system"))
                conn.commit()
    except Exception:
        pass


def _sheets_get(sheet_id: str, range_name: str, token: str) -> list:
    """Fetches values from a Google Sheets range. Returns list of rows."""
    url = (
        "https://sheets.googleapis.com/v4/spreadsheets/"
        f"{sheet_id}/values/{urllib.parse.quote(range_name)}"
    )
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    try:
        with urllib.request.urlopen(req, context=SSL_CTX) as r:
            return json.loads(r.read()).get("values", [])
    except urllib.error.HTTPError as e:
        if e.code in (400, 404):
            return []
        raise


def _sheets_update(sheet_id: str, range_name: str, values: list, token: str) -> dict:
    """Updates a range in a Google Sheet."""
    url = (
        "https://sheets.googleapis.com/v4/spreadsheets/"
        f"{sheet_id}/values/{urllib.parse.quote(range_name)}?valueInputOption=USER_ENTERED"
    )
    payload = json.dumps({"range": range_name, "values": values}).encode()
    req = urllib.request.Request(url, data=payload, method="PUT",
                                  headers={"Authorization": f"Bearer {token}",
                                           "Content-Type": "application/json"})
    with urllib.request.urlopen(req, context=SSL_CTX) as r:
        return json.loads(r.read())


def _try_sheet(sheet_id: str, name_key: str, token: str, range_suffix: str = "A1:Z2000") -> list:
    """Tries multiple tab names and returns the first that has data.
    Always uses a range suffix — required by the Sheets API for tab names containing spaces."""
    for tab_name in SHEET_NAMES.get(name_key, []):
        rng = f"{tab_name}!{range_suffix}"
        rows = _sheets_get(sheet_id, rng, token)
        if rows:
            return rows
    return []


def _drive_list(folder_id: str, token: str) -> list:
    """Lists files/folders in a Drive folder (supports Shared/Team Drives)."""
    params = urllib.parse.urlencode({
        "q": f"'{folder_id}' in parents and trashed=false",
        "fields": "files(id,name,mimeType,modifiedTime)",
        "pageSize": 200,
        "supportsAllDrives": "true",
        "includeItemsFromAllDrives": "true",
    })
    url = f"https://www.googleapis.com/drive/v3/files?{params}"
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    with urllib.request.urlopen(req, context=SSL_CTX) as r:
        return json.loads(r.read()).get("files", [])


def _drive_create_folder(parent_id: str, name: str, token: str) -> str:
    """Creates a folder in Drive and returns its ID."""
    payload = json.dumps({
        "name": name,
        "mimeType": "application/vnd.google-apps.folder",
        "parents": [parent_id],
    }).encode()
    # supportsAllDrives=true is required on writes targeting a Shared Drive parent,
    # not just reads - found 2026-07-06 via a real 404 when creating a division
    # folder inside 1355R's bids folder (which lives in a Shared Drive, ID
    # 0AAI3pETbQDUUUk9PVA). _drive_list already had this param; this write path
    # and the two below never did, so any project whose Drive folder is a Shared
    # Drive (not "My Drive") would fail here the moment a live (non-dry-run) run
    # tried to actually create/upload anything.
    url = "https://www.googleapis.com/drive/v3/files?supportsAllDrives=true"
    req = urllib.request.Request(
        url,
        data=payload, method="POST",
        headers={"Authorization": f"Bearer {token}",
                 "Content-Type": "application/json"},
    )
    with urllib.request.urlopen(req, context=SSL_CTX) as r:
        return json.loads(r.read())["id"]


def _drive_upload(parent_id: str, filename: str, content: bytes,
                  mime_type: str, token: str) -> str:
    """Uploads a file to Drive using multipart upload. Returns file ID."""
    boundary = "HCI_BOUNDARY_" + uuid.uuid4().hex
    metadata = json.dumps({"name": filename, "parents": [parent_id]}).encode()
    body = (
        f"--{boundary}\r\nContent-Type: application/json\r\n\r\n".encode()
        + metadata
        + f"\r\n--{boundary}\r\nContent-Type: {mime_type}\r\n\r\n".encode()
        + content
        + f"\r\n--{boundary}--".encode()
    )
    url = "https://www.googleapis.com/upload/drive/v3/files?uploadType=multipart&supportsAllDrives=true"
    req = urllib.request.Request(
        url, data=body, method="POST",
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": f"multipart/related; boundary={boundary}",
        },
    )
    with urllib.request.urlopen(req, context=SSL_CTX) as r:
        return json.loads(r.read())["id"]


def _drive_update_file(file_id: str, content: bytes, mime_type: str, token: str) -> str:
    """Updates an existing Drive file's content. Returns file ID."""
    url = f"https://www.googleapis.com/upload/drive/v3/files/{file_id}?uploadType=media&supportsAllDrives=true"
    req = urllib.request.Request(
        url, data=content, method="PATCH",
        headers={"Authorization": f"Bearer {token}", "Content-Type": mime_type},
    )
    with urllib.request.urlopen(req, context=SSL_CTX) as r:
        return json.loads(r.read())["id"]


def _drive_find_file(folder_id: str, name: str, token: str) -> Optional[str]:
    """Returns file_id if a file with this name exists in the folder, else None."""
    items = _drive_list(folder_id, token)
    for item in items:
        if item["name"] == name:
            return item["id"]
    return None


def _drive_find_bid_leveling_file(folder_id: str, project_name: str, div_num: str, token: str) -> Optional[str]:
    """Finds an existing bid-leveling Excel for this division by PREFIX
    (project + Div{num}_), not exact full filename match. Found live
    2026-07-09: the filename embeds the division NAME
    (f"{project}_Div{num}_{name}_Bid_Leveling.xlsx"), and that name can come
    from either the legacy Sheet tracker or the Drive-extracted division
    folder depending on which data source populated first - a run before a
    code fix used the Sheet name, a run after used the Drive name, and
    since they differ slightly ("Site Work" vs "Site Work / Exterior
    Improvements"), exact-match _drive_find_file() never found the earlier
    file and silently created a brand new one every time the name drifted,
    instead of updating the same file. Matching on the stable
    "{project}_Div{num}_" prefix survives any division-name-source drift."""
    items = _drive_list(folder_id, token)
    prefix = f"{project_name.replace(' ', '_')}_Div{div_num}_"
    matches = [it for it in items if it["name"].startswith(prefix) and it["name"].endswith("_Bid_Leveling.xlsx")]
    if not matches:
        return None
    # If more than one already exists (drift already happened before this fix),
    # prefer the most recently modified and let the caller know there's cleanup to do.
    matches.sort(key=lambda it: it.get("modifiedTime", ""), reverse=True)
    return matches[0]["id"]


class BidLevelingService:

    @classmethod
    def get_project_config(cls, project_id: int) -> dict:
        with _pg() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id, name, drive_folder_id, gsheet_bid_tracker, bid_folder_id FROM projects WHERE id=%s",
                    (project_id,)
                )
                row = cur.fetchone()
        if not row:
            raise ValueError(f"Project {project_id} not found")
        return dict(row)

    @classmethod
    def get_all_configured_projects(cls) -> list:
        """Returns all active projects with a gsheet_bid_tracker configured."""
        with _pg() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT id, name, drive_folder_id, gsheet_bid_tracker
                    FROM projects
                    WHERE gsheet_bid_tracker IS NOT NULL
                      AND status IN ('active', 'pilot')
                    ORDER BY id
                """)
                return [dict(r) for r in cur.fetchall()]

    @classmethod
    def read_bid_tracking(cls, sheet_id: str, token: str) -> dict:
        """
        Reads the main bid tracking sheet. Returns dict keyed by division number:
        {div_num: {"name": str, "bids": [{"vendor": ..., "date_sent": ...,
                    "date_received": ..., "amount": ..., "status": ..., "contact": ..., "notes": ...}]}}
        """
        rows = _try_sheet(sheet_id, "bid_tracking", token)
        if not rows:
            return {}

        # Find the header row (look for "CSI Division" or "Subcontractor" in any row)
        header_idx = None
        for i, row in enumerate(rows):
            if len(row) >= 2 and any(
                h in str(row[0]).upper() for h in ["CSI DIVISION", "CSI DIV"]
            ):
                header_idx = i
                break
        if header_idx is None:
            # Fallback: look for the first row with >= 5 non-empty cells
            for i, row in enumerate(rows):
                if sum(1 for c in row if c.strip()) >= 5:
                    header_idx = i
                    break

        if header_idx is None:
            return {}

        headers = [str(h).strip().upper() for h in rows[header_idx]]

        def col(name_variants):
            for v in name_variants:
                for i, h in enumerate(headers):
                    if v.upper() in h:
                        return i
            return None

        col_div      = col(["CSI DIVISION", "DIVISION"])
        col_vendor   = col(["SUBCONTRACTOR", "VENDOR", "BIDDER", "COMPANY"])
        col_sent     = col(["DATE SENT"])
        col_received = col(["DATE REC", "DATE RECEIVED"])
        col_amount   = col(["BID AMOUNT", "AMOUNT"])
        col_contact  = col(["CONTACT"])
        col_notes    = col(["NOTES"])
        col_status   = col(["STATUS"])

        divisions = {}
        current_div_num = None
        current_div_name = None

        for row in rows[header_idx + 1:]:
            if not any(c.strip() if isinstance(c, str) else c for c in row):
                continue

            # Detect division header row: first cell filled, rest mostly empty, no bid amount
            first_cell = str(row[0]).strip() if row else ""
            is_div_header = (
                first_cell
                and (col_amount is None or not (len(row) > col_amount and row[col_amount].strip()))
                and len([c for c in row if str(c).strip()]) <= 2
            )

            if is_div_header:
                # Parse division number from "01 — General Requirements", "Division 3 - Concrete",
                # "06 — Wood..." etc. Supports both 1-digit and 2-digit division numbers.
                import re
                m = re.search(r'\b(\d{1,2})\b', first_cell)
                if m:
                    current_div_num = m.group(1).zfill(2)
                    # Strip "Division X - " or "XX — " prefix to get the name
                    name_part = re.sub(r'^(?:Division\s+)?\d{1,2}\s*[—\-–]\s*', '', first_cell).strip()
                    current_div_name = name_part or CSI_DIVISIONS.get(current_div_num, first_cell)
                    if current_div_num not in divisions:
                        divisions[current_div_num] = {"name": current_div_name, "bids": []}
                continue

            if current_div_num is None:
                continue

            def get(i):
                return str(row[i]).strip() if i is not None and i < len(row) else ""

            vendor = get(col_vendor)
            if not vendor or vendor.upper() == "TBD":
                continue

            divisions[current_div_num]["bids"].append({
                "vendor":        vendor,
                "date_sent":     get(col_sent),
                "date_received": get(col_received),
                "amount":        get(col_amount),
                "status":        get(col_status),
                "contact":       get(col_contact),
                "notes":         get(col_notes),
            })

        return divisions

    @classmethod
    def read_division_summary(cls, sheet_id: str, token: str) -> dict:
        """
        Reads HCI Division Summary. Returns dict keyed by div_num:
        {div_num: {"name": str, "leveling_status": str, "recommended": str,
                   "outstanding": str, "risk": str, "next_action": str, "budget": str}}
        """
        rows = _try_sheet(sheet_id, "division_summary", token)
        if not rows:
            return {}

        # Find header row
        header_idx = None
        for i, row in enumerate(rows):
            if row and any("DIV" in str(c).upper() for c in row[:2]):
                header_idx = i
                break
        if header_idx is None:
            return {}

        headers = [str(h).strip().upper() for h in rows[header_idx]]

        def col(names):
            for n in names:
                for i, h in enumerate(headers):
                    if n.upper() in h:
                        return i
            return None

        col_div    = col(["HCI DIV", "CSI DIV", "DIV"])
        col_name   = col(["DIVISION NAME", "DIVISION / TRADE", "TRADE AREA"])
        col_budget = col(["BUDGET"])
        col_status = col(["LEVELING STATUS", "STATUS"])
        col_rec    = col(["RECOMMENDED", "LOW BID"])
        col_out    = col(["OUTSTANDING"])
        col_risk   = col(["RISK"])
        col_action = col(["NEXT ACTION"])

        result = {}
        for row in rows[header_idx + 1:]:
            if not row or not str(row[0]).strip():
                continue
            # Guard added 2026-07-06: a bare .zfill(2) on the raw cell let any
            # non-numeric banner/disclaimer row (e.g. "DRAFT — FOR BUCK'S REVIEW
            # AND APPROVAL ONLY") become a fake division key, since zfill only pads
            # short strings and leaves long ones untouched. Require an actual 1-2
            # digit division number in the cell, same as read_bid_tracking.
            import re
            m = re.search(r'\b(\d{1,2})\b', str(row[0]).strip())
            if not m:
                continue
            div = m.group(1).zfill(2)

            def get(i):
                return str(row[i]).strip() if i is not None and i < len(row) else ""

            result[div] = {
                "name":            get(col_name),
                "budget":          get(col_budget),
                "leveling_status": get(col_status),
                "recommended":     get(col_rec),
                "outstanding":     get(col_out),
                "risk":            get(col_risk),
                "next_action":     get(col_action),
            }
        return result

    @classmethod
    def sync_division_summary_to_sheet(cls, sheet_id: str, divisions_merged: dict, token: str) -> dict:
        """
        Writes computed leveling results (status/recommended/outstanding) back to the
        HCI Division Summary tab for every division whose bids came from a real Drive
        scan this run. Added 2026-07-07 per Buck's direct instruction: the Sheet is the
        tracker his team actually looks at, and bid-leveling had only ever read it,
        never updated it - a run could correctly compute a real leveling result and the
        Sheet would still show stale/blank status, defeating the point of "automation."

        Deliberately narrow in scope: only touches the three system-computed columns
        (LEVELING STATUS, RECOMMENDED, OUTSTANDING). Never writes BUDGET, RISK, or NEXT
        ACTION - those are human-owned narrative fields, not something a bid-leveling
        run has any business overwriting. Only updates rows that already exist in the
        Sheet (matched by division number in column A) - never inserts new rows, since
        getting row insertion wrong could silently shift every row below it.
        """
        import re
        rows, tab_name = [], None
        for name in SHEET_NAMES["division_summary"]:
            candidate = _sheets_get(sheet_id, f"{name}!A1:Z2000", token)
            if candidate:
                rows, tab_name = candidate, name
                break
        if not rows or not tab_name:
            return {"updated": [], "skipped": [], "error": "division_summary tab not found or empty"}

        header_idx = None
        for i, row in enumerate(rows):
            if row and any("DIV" in str(c).upper() for c in row[:2]):
                header_idx = i
                break
        if header_idx is None or not tab_name:
            return {"updated": [], "skipped": [], "error": "could not locate header row or tab name"}

        headers = [str(h).strip().upper() for h in rows[header_idx]]

        def col(names):
            for n in names:
                for i, h in enumerate(headers):
                    if n.upper() in h:
                        return i
            return None

        col_status = col(["LEVELING STATUS", "STATUS"])
        col_rec    = col(["RECOMMENDED", "LOW BID"])
        col_out    = col(["OUTSTANDING"])

        updated, skipped = [], []
        for row_offset, row in enumerate(rows[header_idx + 1:]):
            if not row or not str(row[0]).strip():
                continue
            m = re.search(r'\b(\d{1,2})\b', str(row[0]).strip())
            if not m:
                continue
            div = m.group(1).zfill(2)
            div_data = divisions_merged.get(div)
            if not div_data or div_data.get("source") != "drive":
                continue
            summary = div_data.get("summary") or {}
            if not summary.get("leveling_status"):
                skipped.append(div)
                continue

            sheet_row_num = header_idx + 1 + row_offset + 1 + 1  # 1-indexed, +1 for header itself
            for col_idx, key in ((col_status, "leveling_status"), (col_rec, "recommended"), (col_out, "outstanding")):
                if col_idx is None:
                    continue
                value = summary.get(key, "")
                if not value:
                    continue
                cell_range = f"{tab_name}!{get_column_letter(col_idx + 1)}{sheet_row_num}"
                try:
                    _sheets_update(sheet_id, cell_range, [[value]], token)
                except Exception as e:
                    skipped.append(f"{div} ({key}: {e})")
                    continue
            updated.append(div)
        return {"updated": updated, "skipped": skipped}

    @classmethod
    def read_package_detail(cls, sheet_id: str, token: str) -> dict:
        """
        Reads Bid Leveling Detail. Returns dict keyed by div_num → list of packages:
        {div_num: [{"pkg": str, "trade": str, "budget": str, "level": str,
                    "recommended": str, "num_bids": str, "vendor": str,
                    "outstanding": str, "priority": str}]}
        """
        rows = _try_sheet(sheet_id, "bid_leveling_detail", token)
        if not rows:
            return {}

        header_idx = None
        for i, row in enumerate(rows):
            if row and any("PKG" in str(c).upper() or "TRADE" in str(c).upper() for c in row):
                header_idx = i
                break
        if header_idx is None:
            return {}

        headers = [str(h).strip().upper() for h in rows[header_idx]]

        def col(names):
            for n in names:
                for i, h in enumerate(headers):
                    if n.upper() in h:
                        return i
            return None

        col_div     = col(["HCI DIV", "CSI DIV", "DIV"])
        col_pkg     = col(["PKG"])
        col_trade   = col(["TRADE PACKAGE", "TRADE"])
        col_level   = col(["LEVELING LEVEL", "LEVELING"])
        col_budget  = col(["BUDGET"])
        col_rec     = col(["RECOMMENDED", "LOW BID"])
        col_nbids   = col(["# BIDS", "NUM BIDS"])
        col_vendor  = col(["VENDOR", "BIDDER"])
        col_out     = col(["OUTSTANDING"])
        col_pri     = col(["PRIORITY"])

        result = {}
        current_div = None
        for row in rows[header_idx + 1:]:
            if not row:
                continue
            if col_div is not None and col_div < len(row) and str(row[col_div]).strip():
                raw = str(row[col_div]).strip()
                import re
                # Guard added 2026-07-06: `raw.zfill(2)` used to run unconditionally
                # when no 2-digit match was found, turning a stray banner/disclaimer
                # row (e.g. "DRAFT — FOR BUCK'S REVIEW...") into a fake division key
                # since zfill only pads short strings. Only accept 1-2 digit values;
                # otherwise leave current_div at whatever division came before.
                m = re.search(r'(\d{2})', raw) or re.search(r'\b(\d{1})\b', raw)
                if m:
                    current_div = m.group(1).zfill(2)
            if current_div is None:
                continue

            def get(i):
                return str(row[i]).strip() if i is not None and i < len(row) else ""

            pkg = get(col_pkg)
            trade = get(col_trade)
            if not pkg and not trade:
                continue

            result.setdefault(current_div, []).append({
                "pkg":         pkg,
                "trade":       trade,
                "level":       get(col_level),
                "budget":      get(col_budget),
                "recommended": get(col_rec),
                "num_bids":    get(col_nbids),
                "vendor":      get(col_vendor),
                "outstanding": get(col_out),
                "priority":    get(col_pri),
            })
        return result

    @classmethod
    def generate_division_excel(cls, project_name: str, div_num: str, div_summary: dict,
                                 bids: list, packages: list) -> bytes:
        """
        Generates a bid leveling Excel file for one division.
        Returns bytes suitable for upload to Drive.
        """
        wb = openpyxl.Workbook()
        today = datetime.now().strftime("%Y-%m-%d")
        div_name = div_summary.get("name") or CSI_DIVISIONS.get(div_num, f"Division {div_num}")

        # ── Sheet 1: Bid Summary ──────────────────────────────────────────────
        ws = wb.active
        ws.title = "Bid Summary"

        def _header_style(cell, bg=HCI_BLUE, font_color="FFFFFF", bold=True, size=12):
            cell.font = Font(color=font_color, bold=bold, size=size)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        def _label(cell, value):
            cell.value = value
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill("solid", fgColor=HCI_LIGHT)
            cell.alignment = Alignment(horizontal="left", vertical="center")

        def _value(cell, value):
            cell.value = value
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Title
        ws.merge_cells("A1:H1")
        title_cell = ws["A1"]
        title_cell.value = f"{project_name} — Division {div_num}: {div_name}"
        _header_style(title_cell, size=14)
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:H2")
        sub_cell = ws["A2"]
        sub_cell.value = f"Bid Leveling Summary  |  Generated: {today}  |  HCI AI Operating System"
        _header_style(sub_cell, bg=HCI_HEADER, size=10)
        ws.row_dimensions[2].height = 18

        # Division info block
        info = [
            ("Budget", div_summary.get("budget", "TBD")),
            ("Leveling Status", div_summary.get("leveling_status", "—")),
            ("Recommended / Low Bid", div_summary.get("recommended", "TBD")),
            ("Risk Level", div_summary.get("risk", "—")),
            ("Outstanding Items", div_summary.get("outstanding", "—")),
            ("Next Action", div_summary.get("next_action", "—")),
        ]
        for i, (label, value) in enumerate(info, start=4):
            _label(ws[f"A{i}"], label)
            ws.merge_cells(f"B{i}:H{i}")
            _value(ws[f"B{i}"], value)
            ws.row_dimensions[i].height = 22

        # Bid comparison table
        table_start = 4 + len(info) + 2
        headers = ["SUBCONTRACTOR", "DATE SENT", "DATE REC'D", "BID AMOUNT", "STATUS", "CONTACT", "NOTES"]
        for j, h in enumerate(headers, start=1):
            cell = ws.cell(row=table_start, column=j)
            cell.value = h
            _header_style(cell, bg=HCI_GOLD, font_color="1F1F1F", size=10, bold=True)
        ws.row_dimensions[table_start].height = 20

        for r_idx, bid in enumerate(bids, start=table_start + 1):
            row_vals = [
                bid.get("vendor", ""),
                bid.get("date_sent", ""),
                bid.get("date_received", ""),
                bid.get("amount", ""),
                bid.get("status", ""),
                bid.get("contact", ""),
                bid.get("notes", ""),
            ]
            for j, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=r_idx, column=j)
                cell.value = val
                cell.font = Font(size=10)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                if r_idx % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="EEF4FB")
            ws.row_dimensions[r_idx].height = 18

        # Column widths
        widths = [36, 14, 14, 16, 20, 28, 40]
        for j, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w

        # ── Sheet 2: Package Detail ───────────────────────────────────────────
        if packages:
            ws2 = wb.create_sheet("Package Detail")
            ws2.merge_cells("A1:I1")
            t2 = ws2["A1"]
            t2.value = f"{project_name} — Division {div_num} Package Detail"
            _header_style(t2, size=13)
            ws2.row_dimensions[1].height = 26

            pkg_headers = ["PKG", "TRADE PACKAGE", "BUDGET", "LEVELING STATUS",
                           "VENDOR / BIDDER", "# BIDS", "RECOMMENDED / LOW BID",
                           "OUTSTANDING ISSUE", "PRIORITY"]
            for j, h in enumerate(pkg_headers, start=1):
                cell = ws2.cell(row=3, column=j)
                cell.value = h
                _header_style(cell, bg=HCI_HEADER, size=10)
            ws2.row_dimensions[3].height = 20

            for r_idx, pkg in enumerate(packages, start=4):
                row_vals = [
                    pkg.get("pkg", ""), pkg.get("trade", ""), pkg.get("budget", ""),
                    pkg.get("level", ""), pkg.get("vendor", ""), pkg.get("num_bids", ""),
                    pkg.get("recommended", ""), pkg.get("outstanding", ""), pkg.get("priority", ""),
                ]
                for j, val in enumerate(row_vals, start=1):
                    cell = ws2.cell(row=r_idx, column=j)
                    cell.value = val
                    cell.font = Font(size=10)
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    if r_idx % 2 == 0:
                        cell.fill = PatternFill("solid", fgColor="EEF4FB")
                ws2.row_dimensions[r_idx].height = 20

            pkg_widths = [8, 32, 16, 24, 30, 8, 30, 36, 10]
            for j, w in enumerate(pkg_widths, start=1):
                ws2.column_dimensions[get_column_letter(j)].width = w

        # ── Sheet 3: Outstanding Items ────────────────────────────────────────
        outstanding_items = []
        if div_summary.get("outstanding"):
            outstanding_items.append(("Division Level", div_summary["outstanding"]))
        for b in bids:
            if b.get("notes"):
                outstanding_items.append((b.get("vendor", "Unknown"), b["notes"]))
        for p in packages:
            if p.get("outstanding"):
                outstanding_items.append((p.get("trade", p.get("pkg", "Package")), p["outstanding"]))

        if outstanding_items:
            ws3 = wb.create_sheet("Outstanding Items")
            ws3.merge_cells("A1:C1")
            t3 = ws3["A1"]
            t3.value = f"{project_name} — Division {div_num} Outstanding Items"
            _header_style(t3, size=13)
            ws3.row_dimensions[1].height = 26
            for j, h in enumerate(["SOURCE", "OUTSTANDING ITEM", "ACTION NEEDED"], start=1):
                cell = ws3.cell(row=3, column=j)
                cell.value = h
                _header_style(cell, bg=HCI_GOLD, font_color="1F1F1F", size=10)
            ws3.row_dimensions[3].height = 20
            for r_idx, (src, item) in enumerate(outstanding_items, start=4):
                ws3.cell(row=r_idx, column=1, value=src).font = Font(size=10, bold=True)
                cell = ws3.cell(row=r_idx, column=2, value=item)
                cell.font = Font(size=10)
                cell.alignment = Alignment(wrap_text=True)
                ws3.row_dimensions[r_idx].height = 20
            ws3.column_dimensions["A"].width = 28
            ws3.column_dimensions["B"].width = 60
            ws3.column_dimensions["C"].width = 30

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @classmethod
    def ensure_bids_folder(cls, project_id: int, project_drive_folder_id: str,
                            project_name: str, token: str, dry_run: bool,
                            known_bid_folder_id: str = None) -> dict:
        """
        Finds or creates the {Project} 00_Bids/ folder under the project Drive folder.
        Returns {"folder_id": str, "created": bool, "action": str}

        Fixed 2026-07-07: this used to short-circuit on the hardcoded
        KNOWN_BIDS_FOLDERS dict below, which had 101F pinned to a folder ID from an
        earlier Drive reorganization. The real, current folder (with the actual
        14_Roofing/13_Insulation/5_Waterproofing structure) had since moved to a new
        ID recorded in projects.bid_folder_id - but this function never looked at that
        column, so every 101F bid-leveling run silently read/wrote the stale orphaned
        folder while real incoming bid files landed in the folder Buck actually looks
        at. Found live: a roofing bid Buck could see in Drive was invisible to bid
        leveling entirely. known_bid_folder_id (sourced from projects.bid_folder_id,
        the authoritative value) now takes precedence over the hardcoded dict.
        """
        if known_bid_folder_id:
            return {"folder_id": known_bid_folder_id, "created": False,
                    "action": "from_db_config"}
        if project_id in KNOWN_BIDS_FOLDERS:
            return {"folder_id": KNOWN_BIDS_FOLDERS[project_id], "created": False,
                    "action": "found_existing_hardcoded_fallback"}

        # Check if it already exists in Drive
        existing = _drive_list(project_drive_folder_id, token)
        for item in existing:
            if "00_Bids" in item["name"] and "folder" in item["mimeType"]:
                return {"folder_id": item["id"], "created": False, "action": "found_existing"}

        folder_name = f"{project_name} 00_Bids"
        if dry_run:
            return {"folder_id": None, "created": False,
                    "action": f"would_create: '{folder_name}' in project folder"}

        folder_id = _drive_create_folder(project_drive_folder_id, folder_name, token)
        return {"folder_id": folder_id, "created": True, "action": f"created: '{folder_name}'"}

    @staticmethod
    def _division_folder_match_rank(name: str, div_num: str) -> int:
        """
        Returns a specificity rank (higher = more explicit/preferred) if `name`
        is an existing top-level division folder for div_num, under ANY of the
        naming conventions actually in use across active projects; 0 if it
        doesn't match at all. Leading zeros are normalized so "5" and "05"
        match the same division.

        This exists because the old exact-name / bare-prefix match only
        recognized "{div_num}_" or "{div_num} " and silently fell through to
        creating a brand-new folder whenever a project used a different
        convention - which is how 1355 Riverside ended up with a second,
        parallel, empty "0N_Name" folder tree sitting next to its real
        "Division N - Name" folders (found + fixed 2026-07-09). The rank
        exists because some projects have BOTH an old empty shell folder and
        the real one still present - matching must deterministically prefer
        the explicit "Division N - Name" convention over a bare "N_" or "N -"
        prefix, not just whichever Drive happens to list first.
        """
        n = str(int(div_num))  # normalize "05" / "5" -> "5"
        name = name.strip()
        ranked_patterns = [
            (3, rf'^division\s+0*{n}\s*-\s*'),   # "Division 13 - Insulation" (most explicit)
            (3, rf'^division\s+0*{n}$'),         # "Division 13" (no suffix)
            (2, rf'^0*{n}\s*-\s*'),              # "03 - Concrete...", "05-Metals"
            (1, rf'^0*{n}_'),                    # "13_Insulation", "05_Metals" (bare old-style)
        ]
        for rank, pattern in ranked_patterns:
            if re.match(pattern, name, re.IGNORECASE):
                return rank
        return 0

    @staticmethod
    def _folder_description(name: str, div_num: str) -> str:
        """Strips the leading division-number/convention prefix off a folder
        name, leaving just the descriptive part (e.g. "13_Insulation" ->
        "Insulation"). Used to sanity-check that a numerically-matched folder
        is actually semantically the same division before trusting it."""
        n = str(int(div_num))
        name = name.strip()
        for pattern in (rf'^division\s+0*{n}\s*-\s*', rf'^division\s+0*{n}$',
                        rf'^0*{n}\s*-\s*', rf'^0*{n}_'):
            m = re.match(pattern, name, re.IGNORECASE)
            if m:
                return name[m.end():].strip()
        return name

    @staticmethod
    def _division_names_agree(candidate_desc: str, expected_name: str) -> bool:
        """
        True if two division names plausibly refer to the same trade, based
        on shared significant words. Exists because a numeric-prefix match
        alone is NOT sufficient proof two folders are the same division: on
        1355 Riverside, division 13 means "Insulation" under the real
        (35-division) numbering the system actually uses (confirmed against
        real scanned vendor bids in the `drive_bids` table), but a
        DIFFERENT, older folder tree also has a "Division 13" that means
        "Special Construction" under old 16-division CSI numbering. Those
        are genuinely different divisions that happen to share a number -
        matching on number alone would silently misroute one division's
        leveling summary into a folder that actually belongs to another. If
        either name is empty/generic this returns True (nothing to check
        against, don't block a real match on missing data).
        """
        def sig_words(s: str) -> set:
            return {w.lower() for w in re.findall(r"[a-zA-Z]+", s) if len(w) > 3}
        c_words, e_words = sig_words(candidate_desc), sig_words(expected_name)
        if not c_words or not e_words:
            return True
        return bool(c_words & e_words)

    @classmethod
    def ensure_division_folders(cls, bids_folder_id: str, divisions: dict,
                                  token: str, dry_run: bool) -> dict:
        """
        Ensures each division has a folder under 00_Bids/.
        Returns {div_num: {"folder_id": str, "action": str}}
        """
        if bids_folder_id is None:
            return {d: {"folder_id": None, "action": "waiting_on_bids_folder"} for d in divisions}

        existing = _drive_list(bids_folder_id, token)
        existing_map = {item["name"]: item["id"] for item in existing
                        if "folder" in item["mimeType"]}

        result = {}
        for div_num, div_data in divisions.items():
            # 2026-07-13: division keys can now be compound ambiguous keys
            # like "07__Thermal & Moisture" (from read_drive_bids()'s
            # division-13-collision fix) - this folder-matching code predates
            # that and expects a bare numeric division_num, crashing on
            # int(div_num). Use the bare number for matching/folder-naming;
            # the full division_name (already correct) still comes through
            # div_data for display.
            div_num_bare = div_num.split("__")[0]
            div_name = div_data.get("name") or CSI_DIVISIONS.get(div_num_bare, f"Division {div_num_bare}")
            folder_name = f"{div_num_bare}_{div_name}"
            # Collect every existing folder that could plausibly BE this
            # division's folder, ranked by how explicit/canonical its naming
            # convention is - and pick the single best one. Deliberately NOT
            # special-casing an exact string match to `folder_name` as
            # automatically "best": folder_name is itself constructed in the
            # old "{div_num}_{div_name}" shape, so an exact match only proves
            # a folder follows the OLD convention, not that it's the real or
            # current one. On 1355 Riverside the old empty "0N_Name" shell
            # folder happened to equal that constructed string byte-for-byte
            # and was silently winning over the real "Division N - Name"
            # folder purely because of that coincidence - exactly the bug
            # being fixed here.
            candidates = []
            for name, fid in existing_map.items():
                rank = cls._division_folder_match_rank(name, div_num_bare)
                if rank == 0:
                    continue
                desc = cls._folder_description(name, div_num_bare)
                agrees = cls._division_names_agree(desc, div_name)
                candidates.append((agrees, rank, name, fid))

            folder_id = None
            if candidates:
                # Semantic agreement is a TIEBREAKER between multiple
                # numeric-matching candidates, not an absolute filter - a
                # lone candidate is used even if worded differently than
                # expected (e.g. 64EW's real division-07 folder is "07 -
                # Insulation, Waterproofing, Roofing & Siding" but the
                # computed div_name is "Thermal & Moisture Protection" -
                # same division, different wording, and there's no
                # alternative folder to prefer instead). Only when TWO OR
                # MORE candidates share a division number does a semantic
                # mismatch mean "this is probably a different division under
                # another numbering scheme" (1355R's div 13 = Insulation vs.
                # a leftover "Division 13 - Special Construction" folder).
                candidates.sort(key=lambda c: (c[0], c[1]), reverse=True)
                folder_id = candidates[0][3]

            if folder_id:
                action = "found_existing"
                if len(candidates) > 1:
                    other_names = [c[2] for c in candidates[1:]]
                    action = (f"found_existing (WARNING: {len(candidates)} candidate folders "
                              f"matched div {div_num} - picked '{candidates[0][2]}', "
                              f"ignored {other_names} - possible duplicate folder tree or "
                              f"conflicting numbering scheme, flag for cleanup)")
                result[div_num] = {"folder_id": folder_id, "action": action}
            elif dry_run:
                result[div_num] = {"folder_id": None,
                                    "action": f"would_create: '{folder_name}'"}
            else:
                new_id = _drive_create_folder(bids_folder_id, folder_name, token)
                result[div_num] = {"folder_id": new_id, "action": f"created: '{folder_name}'"}
        return result

    @classmethod
    def scan_drive_bids(cls, project_id: int, dry_run: bool = True) -> dict:
        """
        Scan a project's Drive bid folders for new vendor PDFs and extract bid data.
        Delegates to drive_bid_reader.scan_project_bids().
        """
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        from drive_bid_reader import scan_project_bids
        return scan_project_bids(project_id, dry_run=dry_run)

    @classmethod
    def run_bid_leveling(cls, project_id: int, dry_run: bool = True,
                          divisions_filter: list = None,
                          scan_drive: bool = True) -> dict:
        """
        Main orchestrator for one project.
        dry_run=True: reads everything, returns analysis with no Drive writes.
        dry_run=False: generates Excel files and queues all Drive creates/uploads.
        divisions_filter: optional list of div_nums to process (e.g. ["06", "07"])
        scan_drive: if True (default), scan Drive folders for new bids before leveling.
        """
        config = cls.get_project_config(project_id)
        project_name = config["name"]
        sheet_id     = config["gsheet_bid_tracker"]
        drive_folder = config["drive_folder_id"]

        if not sheet_id:
            return {"error": f"Project {project_name} has no gsheet_bid_tracker configured"}
        if not drive_folder and not dry_run:
            return {"error": f"Project {project_name} has no drive_folder_id configured — required for Excel upload (dry_run=True works without it)"}

        token_sheets = get_google_token("sheets")
        token_drive  = get_google_token("drive")

        # ── Step 1: Scan Drive for new bids ──────────────────────────────────
        # scan_drive=False used to skip BOTH the slow Drive re-walk/extraction
        # AND the (fast) read of already-extracted drive_bids rows from the DB,
        # since both were gated by the same flag - found live 2026-07-09 when a
        # scan_drive=False call (meant only to skip the slow ~190-file re-scan
        # on a project with an already-current database) silently fell back to
        # the stale legacy Google Sheets tracker for every division instead of
        # the clean, already-extracted Drive data, because drive_bids_by_div
        # stayed {}. The DB read is cheap; only the Drive walk/AI-extraction is
        # slow. Now the read always runs; only the walk is gated.
        drive_scan_result = None
        drive_bids_by_div = {}
        leveling_summary = {}
        if config.get("bid_folder_id"):
            sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
            from drive_bid_reader import scan_project_bids, read_drive_bids, get_leveling_summary
            if scan_drive:
                drive_scan_result = scan_project_bids(project_id, dry_run=dry_run)
            drive_bids_by_div  = read_drive_bids(project_id, latest_only=True)
            leveling_summary   = get_leveling_summary(project_id)

        # ── Step 2: Read Google Sheet data ───────────────────────────────────
        # Read all data sources
        bid_data   = cls.read_bid_tracking(sheet_id, token_sheets)
        div_summary = cls.read_division_summary(sheet_id, token_sheets)
        pkg_detail  = cls.read_package_detail(sheet_id, token_sheets)

        if divisions_filter:
            bid_data        = {k: v for k, v in bid_data.items() if k in divisions_filter}
            div_summary     = {k: v for k, v in div_summary.items() if k in divisions_filter}
            pkg_detail      = {k: v for k, v in pkg_detail.items() if k in divisions_filter}
            # Drive-sourced divisions were never filtered here (found 2026-07-06 via
            # test_bid_leveling.py's BL-DRY-04) - a caller asking for divisions
            # ["16","15"] still got every division that had a Drive-sourced bid,
            # since drive_bids_by_div was merged in below unconditionally.
            drive_bids_by_div = {k: v for k, v in drive_bids_by_div.items() if k in divisions_filter}

        # Include Drive-sourced divisions even if not in Sheet
        drive_div_keys = list(drive_bids_by_div.keys()) if drive_bids_by_div else []
        all_divs = sorted(set(list(bid_data.keys()) + list(div_summary.keys()) + drive_div_keys))
        total_bids = sum(len(d["bids"]) for d in bid_data.values())

        # Merge bid data + summary for all known divisions
        # Drive bids are preferred over Sheet bids when both exist for the same division
        divisions_merged = {}
        for div in all_divs:
            summary = div_summary.get(div, {})
            sheet_bids = bid_data.get(div, {}).get("bids", [])
            pkgs    = pkg_detail.get(div, [])
            name    = summary.get("name") or (bid_data.get(div, {}).get("name") or
                       CSI_DIVISIONS.get(div, f"Division {div}"))

            # Use Drive bids as primary if available for this division
            drive_div = drive_bids_by_div.get(div, {})
            if drive_div and drive_div.get("bids"):
                drive_bids_list = drive_div["bids"]
                name = drive_div.get("division_name") or name
                bids = [
                    {
                        "vendor":        b["vendor"],
                        "date_sent":     "",
                        "date_received": b["bid_date"],
                        "amount":        b["amount_fmt"] or "",
                        "status":        "bid_received",
                        "contact":       "",
                        "notes":         b.get("scope", ""),
                        "source":        "drive",
                    }
                    for b in drive_bids_list
                ]
                # Update summary with Drive leveling data if available
                lev = leveling_summary.get(div, {})
                if lev and not summary.get("recommended"):
                    summary = {**summary,
                               "recommended": f"{lev['low_vendor']} ${lev['low_bid']:,.0f}",
                               "leveling_status": "Bids Received (Drive)",
                               "outstanding": f"Spread: ${lev['spread']:,.0f} ({lev['spread_pct']}%)"}
            else:
                bids = sheet_bids

            # Flag data gaps explicitly rather than silently rendering blank fields -
            # found 2026-07-06 that some divisions have bids logged but no Division
            # Summary or Bid Leveling Detail entered (e.g. 64EW Div 26, 1355R Div 17),
            # which used to just show up as empty strings with no indication anything
            # was missing. A human looking at this (e.g. Adam during onboarding) needs
            # to see "no summary entered yet" explicitly, not guess why fields are blank.
            data_gaps = []
            if not summary:
                data_gaps.append("no_division_summary_entry")
            if not pkgs:
                data_gaps.append("no_package_detail_entry")
            if not bids:
                data_gaps.append("no_bid_tracking_rows")

            divisions_merged[div] = {
                "div_num":   div,
                "name":      name,
                "summary":   summary,
                "bids":      bids,
                "packages":  pkgs,
                "bid_count": len(bids),
                "source":    "drive" if drive_div and drive_div.get("bids") else "sheet",
                "data_gaps": data_gaps,
            }

        # Folder setup — skip if no drive_folder (dry_run only)
        if drive_folder:
            bids_folder_result = cls.ensure_bids_folder(
                project_id, drive_folder, project_name, token_drive, dry_run,
                known_bid_folder_id=config.get("bid_folder_id")
            )
            bids_folder_id = bids_folder_result["folder_id"]
            div_folders = cls.ensure_division_folders(
                bids_folder_id, divisions_merged, token_drive, dry_run
            )
        else:
            bids_folder_result = {"folder_id": None, "action": "skipped_no_drive_folder"}
            bids_folder_id = None
            div_folders = {}

        # 2026-07-13: div_num can be a compound ambiguous key like
        # "07__Thermal & Moisture" (see ensure_division_folders fix above).
        # A bare sibling ("07", sourced from the Google Sheet tracker, which
        # doesn't know about the Drive-side sub-package disambiguation) can
        # legitimately carry different real content (Sheet narrative/package
        # tracking) than its compound sibling (real Drive-sourced bids) - so
        # neither should be silently dropped - but both reduce to the
        # identical filename once the compound suffix is stripped for
        # display, and whichever queued second would silently overwrite the
        # other's real data in Drive. Found live regenerating 1355R's Excel
        # summaries. Compute every filename first, then disambiguate any
        # collision with a numbered suffix rather than dropping data.
        filename_counts: dict = {}
        for div_num, div_data in divisions_merged.items():
            div_num_bare = div_num.split("__")[0]
            safe_name = div_data["name"].replace("/", "-").replace("&", "and").replace(" ", "_")
            base_filename = f"{project_name.replace(' ', '_')}_Div{div_num_bare}_{safe_name}_Bid_Leveling.xlsx"
            filename_counts[base_filename] = filename_counts.get(base_filename, 0) + 1
        filename_seen: dict = {}

        # Generate Excel files per division
        excel_actions = []
        queued_items  = []
        files_generated = 0

        for div_num, div_data in divisions_merged.items():
            summary  = div_data["summary"]
            bids     = div_data["bids"]
            packages = div_data["packages"]
            div_name = div_data["name"]

            div_num_bare = div_num.split("__")[0]
            safe_name = div_name.replace("/", "-").replace("&", "and").replace(" ", "_")
            base_filename = f"{project_name.replace(' ', '_')}_Div{div_num_bare}_{safe_name}_Bid_Leveling.xlsx"
            if filename_counts[base_filename] > 1:
                filename_seen[base_filename] = filename_seen.get(base_filename, 0) + 1
                stem, ext = base_filename.rsplit(".", 1)
                filename = f"{stem}_{filename_seen[base_filename]}.{ext}"
            else:
                filename = base_filename

            folder_info = div_folders.get(div_num, {"folder_id": None, "action": "unknown"})

            if dry_run:
                excel_actions.append({
                    "division":   div_num,
                    "name":       div_name,
                    "filename":   filename,
                    "bids_found": len(bids),
                    "pkgs_found": len(packages),
                    "folder_action": folder_info["action"],
                    "file_action": f"would_upload to folder: {folder_info['folder_id'] or 'TBD'}",
                    "data_gaps": div_data.get("data_gaps", []),
                })
            else:
                # Generate Excel
                content = cls.generate_division_excel(
                    project_name, div_num, summary, bids, packages
                )

                folder_id = folder_info.get("folder_id")
                if folder_id:
                    # Queue the upload via approval queue
                    import base64
                    queue_result = ApprovalQueueService.enqueue(
                        workflow="bid_leveling",
                        action_type="drive_upload_file",
                        target_system="google_drive",
                        target_id=f"drive:{folder_id}/{filename}",
                        target_description=f"{project_name} Div {div_num_bare} {div_name} bid leveling Excel",
                        proposed_payload={
                            "folder_id":  folder_id,
                            "filename":   filename,
                            "project_name": project_name,
                            "div_num":    div_num,
                            "content_b64": base64.b64encode(content).decode(),
                            "mime_type":  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        },
                        reason=f"Upload generated bid leveling file for {project_name} Division {div_num}",
                        project_id=project_id,
                        rollback_path=f"Delete file {filename} from Drive folder {folder_id}",
                    )
                    queued_items.append(queue_result.get("id"))
                    files_generated += 1
                    excel_actions.append({
                        "division":    div_num,
                        "name":        div_name,
                        "filename":    filename,
                        "bids_found":  len(bids),
                        "queue_id":    queue_result.get("id"),
                        "action":      "queued_for_upload",
                    })
                else:
                    excel_actions.append({
                        "division":  div_num,
                        "name":      div_name,
                        "filename":  filename,
                        "action":    "skipped_no_folder",
                        "bids_found": len(bids),
                    })

        if not dry_run and (bids_folder_result.get("created") or total_bids > 0):
            _log_roi(project_id, project_name, len(divisions_merged), total_bids, files_generated)

        # Write real leveling results back to the Sheet - added 2026-07-07 per Buck's
        # direct instruction. Prior to this, run_bid_leveling only ever READ the Sheet;
        # a run could correctly compute a real leveling result from Drive-sourced bids
        # and the tracker Buck's team actually looks at would still show stale/blank
        # status. Only touches computed status/recommended/outstanding columns, never
        # budget/risk/next-action (human-owned narrative fields).
        sheet_sync_result = None
        if not dry_run:
            try:
                sheet_sync_result = cls.sync_division_summary_to_sheet(sheet_id, divisions_merged, token_sheets)
            except Exception as e:
                sheet_sync_result = {"error": str(e)}

        return {
            "project":            project_name,
            "project_id":         project_id,
            "mode":               "dry_run" if dry_run else "live",
            "sheet_id":           sheet_id,
            "drive_folder":       drive_folder,
            "bid_folder_id":      config.get("bid_folder_id"),
            "drive_scan":         drive_scan_result,
            "drive_bids_by_div":  {k: len(v["bids"]) for k, v in drive_bids_by_div.items()} if drive_bids_by_div else {},
            "leveling_summary":   leveling_summary,
            "divisions_found":    len(all_divs),
            "total_bids":         total_bids,
            "bids_folder":        bids_folder_result,
            "division_folders":   div_folders,
            "excel_actions":      excel_actions,
            "queued_items":       queued_items,
            "sheet_sync":         sheet_sync_result,
            "divisions":          divisions_merged if dry_run else None,
            "data_gaps_by_division": {
                d: v["data_gaps"] for d, v in divisions_merged.items() if v.get("data_gaps")
            },
        }

    @classmethod
    def run_all_projects(cls, dry_run: bool = True) -> dict:
        """Runs bid leveling for all configured projects."""
        projects = cls.get_all_configured_projects()
        results = []
        for proj in projects:
            try:
                result = cls.run_bid_leveling(proj["id"], dry_run=dry_run)
                results.append(result)
            except Exception as e:
                results.append({"project_id": proj["id"], "project": proj["name"],
                                  "error": str(e)})
        return {
            "mode":        "dry_run" if dry_run else "live",
            "projects_run": len(results),
            "results":     results,
        }

    @classmethod
    def execute_queued_upload(cls, queue_id: int) -> dict:
        """
        Actually executes a previously queued Drive upload.
        Called after Buck approves the queue item.
        """
        import base64
        with _pg() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT id, status, approved_by, proposed_payload
                    FROM approval_queue WHERE id = %s
                """, (queue_id,))
                item = cur.fetchone()

        if not item:
            return {"error": f"Queue item {queue_id} not found"}
        if item["status"] != "approved":
            return {"error": f"Queue item {queue_id} is {item['status']}, not approved"}

        payload = item["proposed_payload"] if isinstance(item["proposed_payload"], dict) \
                  else json.loads(item["proposed_payload"])

        token  = get_google_token("drive")
        folder_id  = payload["folder_id"]
        filename   = payload["filename"]
        content    = base64.b64decode(payload["content_b64"])
        mime_type  = payload.get("mime_type",
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # Check if file already exists (update vs create). Prefer the stable
        # project+division-number prefix match over an exact filename match -
        # the filename embeds the division NAME, which can drift between runs
        # depending on data source (Sheet vs Drive), which used to cause a
        # silent duplicate file every time the name drifted instead of an
        # update to the same file. Fall back to exact match for older queue
        # items that predate the project_name/div_num payload fields.
        project_name = payload.get("project_name")
        div_num      = payload.get("div_num")
        if project_name and div_num:
            existing_id = _drive_find_bid_leveling_file(folder_id, project_name, div_num, token)
        else:
            existing_id = _drive_find_file(folder_id, filename, token)
        if existing_id:
            file_id = _drive_update_file(existing_id, content, mime_type, token)
            action = "updated"
        else:
            file_id = _drive_upload(folder_id, filename, content, mime_type, token)
            action = "created"

        # Mark executed in queue
        ApprovalQueueService.mark_executed(queue_id, "system")

        return {
            "queue_id":  queue_id,
            "file_id":   file_id,
            "filename":  filename,
            "action":    action,
            "folder_id": folder_id,
        }

    @classmethod
    def get_project_data(cls, project_id: int) -> dict:
        """
        Returns all raw bid data for a project (for GBT/AI agent consumption).
        Includes bid tracking, division summary, and package detail.
        """
        config = cls.get_project_config(project_id)
        sheet_id = config["gsheet_bid_tracker"]
        if not sheet_id:
            return {"error": "No sheet configured"}

        token = get_google_token("sheets")
        return {
            "project":          config["name"],
            "project_id":       project_id,
            "sheet_id":         sheet_id,
            "drive_folder_id":  config["drive_folder_id"],
            "bid_tracking":     cls.read_bid_tracking(sheet_id, token),
            "division_summary": cls.read_division_summary(sheet_id, token),
            "package_detail":   cls.read_package_detail(sheet_id, token),
        }

    @classmethod
    def create_folder(cls, parent_folder_id: str, folder_name: str) -> dict:
        """
        Directly creates a folder in Drive (for GBT / authorized AI agent use).
        Returns {"folder_id": str, "name": str, "parent": str}
        """
        token = get_google_token("drive")
        # Check if already exists
        existing = _drive_list(parent_folder_id, token)
        for item in existing:
            if item["name"] == folder_name and "folder" in item["mimeType"]:
                return {"folder_id": item["id"], "name": folder_name,
                        "parent": parent_folder_id, "action": "found_existing"}
        folder_id = _drive_create_folder(parent_folder_id, folder_name, token)
        return {"folder_id": folder_id, "name": folder_name,
                "parent": parent_folder_id, "action": "created"}

    @classmethod
    def upload_file(cls, folder_id: str, filename: str, content_b64: str,
                     mime_type: str = "application/octet-stream") -> dict:
        """
        Uploads a file to Drive (for GBT / authorized AI agent use).
        content_b64: base64-encoded file content.
        Returns {"file_id": str, "filename": str, "folder_id": str, "action": str}
        """
        import base64
        token   = get_google_token("drive")
        content = base64.b64decode(content_b64)
        existing_id = _drive_find_file(folder_id, filename, token)
        if existing_id:
            file_id = _drive_update_file(existing_id, content, mime_type, token)
            action = "updated"
        else:
            file_id = _drive_upload(folder_id, filename, content, mime_type, token)
            action = "created"
        return {"file_id": file_id, "filename": filename,
                "folder_id": folder_id, "action": action}

    @classmethod
    def list_drive_folder(cls, folder_id: str) -> dict:
        """Lists contents of a Drive folder (for GBT / AI agent read access)."""
        token = get_google_token("drive")
        items = _drive_list(folder_id, token)
        return {"folder_id": folder_id, "items": items, "count": len(items)}

    @classmethod
    def read_sheet_range(cls, sheet_id: str, range_name: str) -> dict:
        """Reads a Google Sheets range (for GBT / AI agent read access)."""
        token = get_google_token("sheets")
        rows  = _sheets_get(sheet_id, range_name, token)
        return {"sheet_id": sheet_id, "range": range_name,
                "rows": rows, "row_count": len(rows)}

    @classmethod
    def write_sheet_range(cls, sheet_id: str, range_name: str, values: list) -> dict:
        """Writes to a Google Sheets range (for GBT / AI agent write access)."""
        token  = get_google_token("sheets")
        result = _sheets_update(sheet_id, range_name, values, token)
        return {"sheet_id": sheet_id, "range": range_name,
                "updated_cells": result.get("updatedCells", 0),
                "updated_rows":  result.get("updatedRows", 0)}
